#!/usr/bin/env python3
"""
audit_club_nicknames.py -- What club identity and nickname data already exists?

Read-only. Writes nothing, changes nothing. Run this before building the
canonical club layer so the plan is based on measured coverage rather than
assumption.

    python audit_club_nicknames.py
    python audit_club_nicknames.py --db data/afl/afl.db
    python audit_club_nicknames.py --xlsx club_info.xlsx     # optional compare
    python audit_club_nicknames.py --csv                     # machine-readable

Answers five questions:

  1. What is in `clubs` already, and is `active` set sensibly?
  2. Which clubs have a scraped nickname in `club_wikipedia_fields`?
  3. Which club identities appear in `games` but not in `clubs`?
  4. Which alias literals scattered through the codebase fail to resolve?
  5. Where does the spreadsheet disagree with the scrape?

Question 3 is the one that matters most. `clubs` is populated from the
club-sources scrape, which covers current clubs; `games.club_hist` carries
Fitzroy, University, Brisbane Bears, South Melbourne, Footscray and Kangaroos.
Anything listed there is an identity the canonical layer must hand-author.
"""

import argparse
import csv
import sqlite3
import sys
from pathlib import Path

# Historical identities that are expected to be absent from `clubs` and must
# be hand-authored. Listing them makes the audit's output actionable: an
# identity that is missing AND not listed here is an unexpected gap.
KNOWN_HISTORICAL = {
    "Fitzroy", "University", "Brisbane Bears", "South Melbourne",
    "Footscray", "Kangaroos",
}

# Nicknames for those identities, for the report only. The script does not
# write these; the loader in step 2 does.
HISTORICAL_NICKNAMES = {
    "Fitzroy": "Lions",
    "University": "Students",
    "Brisbane Bears": "Bears",
    "South Melbourne": "Swans",
    "Footscray": "Bulldogs",
    "Kangaroos": "Kangaroos",
}


def table_exists(con, name):
    return con.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
    ).fetchone() is not None


def columns(con, table):
    return [r[1] for r in con.execute(f"PRAGMA table_info({table})")]


def rule(title):
    print(f"\n{title}\n{'-' * len(title)}")


# ------------------------------------------------------------------ 1. clubs

def audit_clubs_table(con):
    rule("1. clubs table")
    if not table_exists(con, "clubs"):
        print("  clubs: MISSING -- club sources have not been loaded.")
        print("  Run load_club_sources.py before the canonical club layer.")
        return []

    cols = columns(con, "clubs")
    print(f"  columns: {', '.join(cols)}")

    rows = con.execute(
        "SELECT club_id, name, abbreviation, db_club_now, active "
        "FROM clubs ORDER BY name"
    ).fetchall()
    print(f"  rows: {len(rows)}  "
          f"(active={sum(r[4] for r in rows)}, inactive={sum(1 - r[4] for r in rows)})")

    # Columns the canonical layer needs to add.
    wanted = ["nickname", "state", "home_venue", "established", "first_season",
              "active_from", "active_to", "predecessor_club", "successor_club",
              "source"]
    missing = [c for c in wanted if c not in cols]
    if missing:
        print(f"  columns to ADD in step 2: {', '.join(missing)}")
    else:
        print("  canonical columns already present.")

    for club_id, name, abbr, db_now, active in rows:
        flag = "" if active else "  [inactive]"
        print(f"    {club_id:<24} {name:<26} {abbr:<5} -> games.club_now="
              f"{db_now!r}{flag}")
    return rows


# -------------------------------------------------------- 2. scraped nickname

def audit_scraped_nicknames(con, club_rows):
    rule("2. nickname coverage in club_wikipedia_fields")
    if not table_exists(con, "club_wikipedia_fields"):
        print("  club_wikipedia_fields: MISSING.")
        return {}

    found = {}
    for key in ("nickname", "nicknames", "nickname_s"):
        for club_id, value in con.execute(
            "SELECT club_id, field_value FROM club_wikipedia_fields "
            "WHERE field_key=? AND field_value IS NOT NULL "
            "AND TRIM(field_value) != ''", (key,)
        ):
            found.setdefault(club_id, (value, key))

    have, lack = [], []
    for club_id, name, *_ in club_rows:
        (have if club_id in found else lack).append((club_id, name))

    print(f"  {len(have)} of {len(club_rows)} clubs have a scraped nickname.")
    for club_id, name in have:
        value, key = found[club_id]
        via = "" if key == "nickname" else f"   (via field_key={key!r})"
        print(f"    {name:<26} {value}{via}")
    if lack:
        print("  NO scraped nickname:")
        for club_id, name in lack:
            print(f"    {name:<26} ({club_id})")

    # Multi-valued nicknames are common on Wikipedia ("Dogs, Scray, Bulldogs")
    # and must not be written straight into a single canonical column.
    multi = [(cid, v) for cid, (v, _) in found.items()
             if "," in v or " or " in v.lower() or "/" in v]
    if multi:
        print("  MULTI-VALUED (needs a primary/alias split in step 2):")
        for club_id, value in multi:
            print(f"    {club_id:<24} {value}")
    return {cid: v for cid, (v, _) in found.items()}


# ------------------------------------------------ 3. identities used by games

def audit_game_identities(con, club_rows):
    rule("3. club identities in games, not in clubs")
    if not table_exists(con, "games"):
        print("  games: MISSING.")
        return []

    game_cols = columns(con, "games")
    known = {r[3] for r in club_rows}          # db_club_now
    known |= {r[1] for r in club_rows}         # name

    identities = {}
    for col in ("club_now", "club_hist"):
        if col not in game_cols:
            continue
        for value, n in con.execute(
            f"SELECT {col}, COUNT(*) FROM games "
            f"WHERE {col} IS NOT NULL AND TRIM({col}) != '' "
            f"GROUP BY {col} ORDER BY COUNT(*) DESC"
        ):
            entry = identities.setdefault(value, {"club_now": 0, "club_hist": 0})
            entry[col] = n

    unmapped = [(v, c) for v, c in identities.items() if v not in known]
    print(f"  {len(identities)} distinct identities in games; "
          f"{len(unmapped)} not resolvable through clubs.")

    expected, unexpected = [], []
    for value, counts in sorted(unmapped, key=lambda x: -sum(x[1].values())):
        (expected if value in KNOWN_HISTORICAL else unexpected).append(
            (value, counts))

    if expected:
        print("  EXPECTED historical identities (hand-author in step 2):")
        for value, counts in expected:
            nick = HISTORICAL_NICKNAMES.get(value, "?")
            print(f"    {value:<20} now={counts['club_now']:>6,} "
                  f"hist={counts['club_hist']:>6,}   nickname -> {nick}")
    if unexpected:
        print("  UNEXPECTED -- investigate before authoring anything:")
        for value, counts in unexpected:
            print(f"    {value:<20} now={counts['club_now']:>6,} "
                  f"hist={counts['club_hist']:>6,}")
    if not unmapped:
        print("  All game identities resolve.")

    missing_historical = KNOWN_HISTORICAL - set(identities)
    if missing_historical:
        print(f"  NOTE: expected but absent from games entirely: "
              f"{', '.join(sorted(missing_historical))}")
    return unmapped


# ------------------------------------------------------- 4. codebase aliases

def audit_code_aliases(con, club_rows):
    """Do the alias literals in afl/parse_criteria.py all resolve?

    Imported rather than duplicated, so this stays true as the dict changes.
    """
    rule("4. parse_criteria.CLUB_ALIASES resolution")
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from afl.parse_criteria import CLUB_ALIASES
    except Exception as exc:
        print(f"  could not import parse_criteria ({exc}).")
        print("  Run from the repo root, or skip -- this check is advisory.")
        return

    known = {r[1] for r in club_rows} | {r[3] for r in club_rows}
    if table_exists(con, "games"):
        cols = columns(con, "games")
        for col in ("club_now", "club_hist"):
            if col in cols:
                known |= {r[0] for r in con.execute(
                    f"SELECT DISTINCT {col} FROM games WHERE {col} IS NOT NULL")}

    bad = sorted({t for t in CLUB_ALIASES.values() if t not in known})
    print(f"  {len(CLUB_ALIASES)} aliases -> {len(set(CLUB_ALIASES.values()))} "
          f"targets; {len(bad)} targets unknown to the database.")
    for target in bad:
        sources = [a for a, t in CLUB_ALIASES.items() if t == target]
        print(f"    {target!r} <- {', '.join(sorted(sources))}")
    if not bad:
        print("  Every alias target resolves.")


# ----------------------------------------------------- 5. spreadsheet compare

def audit_spreadsheet(path, club_rows, scraped):
    rule("5. club_info.xlsx vs scraped nicknames")
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed -- skipping.")
        return
    if not Path(path).exists():
        print(f"  {path} not found -- skipping.")
        return

    ws = openpyxl.load_workbook(path, data_only=True).active
    sheet = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0] or row[0] in ("Current clubs", "Club"):
            continue
        name, colours, moniker = row[0], row[1], row[2]
        if moniker:
            sheet[str(name).strip()] = str(moniker).strip()
        if colours == "#VALUE!":
            pass  # counted below

    broken = sum(1 for r in ws.iter_rows(values_only=True)
                 if r and r[1] == "#VALUE!")
    print(f"  {len(sheet)} monikers read; colours column broken in "
          f"{broken} rows (#VALUE!) -- do not import that column.")

    by_name = {name: club_id for club_id, name, *_ in club_rows}
    agree = disagree = only_sheet = 0
    for name, moniker in sorted(sheet.items()):
        club_id = by_name.get(name)
        if club_id is None:
            # Spreadsheet uses "Greater Western Sydney"; the db may use "GWS".
            print(f"    {name:<26} sheet={moniker:<12} NO clubs row matched "
                  f"on name -- needs an alias")
            only_sheet += 1
            continue
        current = scraped.get(club_id)
        if current is None:
            print(f"    {name:<26} sheet={moniker:<12} scraped=(none) -> fill")
        elif moniker.lower() in current.lower():
            agree += 1
        else:
            disagree += 1
            print(f"    {name:<26} sheet={moniker:<12} scraped={current} "
                  f"*** DISAGREE -- record both, do not overwrite")
    print(f"  agree={agree}  disagree={disagree}  unmatched_name={only_sheet}")


# ---------------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", default=None,
                    help="database path (default: data_paths.default_db('afl'))")
    ap.add_argument("--xlsx", default="club_info.xlsx")
    ap.add_argument("--csv", action="store_true",
                    help="also write club_nickname_audit.csv")
    args = ap.parse_args()

    db = args.db
    if db is None:
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            from data_paths import default_db
            db = default_db("afl")
        except Exception:
            db = "gridley.db"
    if not Path(db).exists():
        sys.exit(f"database not found: {db}")

    print(f"database: {db}")
    con = sqlite3.connect(f"file:{db}?mode=ro", uri=True)
    try:
        club_rows = audit_clubs_table(con)
        scraped = audit_scraped_nicknames(con, club_rows) if club_rows else {}
        unmapped = audit_game_identities(con, club_rows)
        audit_code_aliases(con, club_rows)
        audit_spreadsheet(args.xlsx, club_rows, scraped)

        rule("summary")
        print(f"  clubs rows              {len(club_rows)}")
        print(f"  with scraped nickname   {len(scraped)}")
        print(f"  game identities unmapped {len(unmapped)}")
        print("  Hand-authored rows needed for every identity listed under 3.")

        if args.csv:
            out = Path("club_nickname_audit.csv")
            with out.open("w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(["club_id", "name", "db_club_now", "active",
                            "scraped_nickname"])
                for club_id, name, abbr, db_now, active in club_rows:
                    w.writerow([club_id, name, db_now, active,
                                scraped.get(club_id, "")])
                for value, counts in unmapped:
                    w.writerow(["", value, value, "",
                                HISTORICAL_NICKNAMES.get(value, "")])
            print(f"  wrote {out}")
    finally:
        con.close()


if __name__ == "__main__":
    main()
